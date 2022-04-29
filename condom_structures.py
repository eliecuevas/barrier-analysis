from typing import Pattern
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles.colors import Color
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

from response_class import Response




import os

datawb = load_workbook("CondomSurvey.xlsx")
data = datawb.active

row = 2
total_responses = 0


#generic_agree_responses = {"Strongly agree", "Agree", "Neither agree nor disagree", "Disagree", "Strongly disagree"}

total_confidence_responses = 0
s_a = "strongly agree"
_a_ = "agree"
n_a_d  = "neither agree nor disagree"
_d_ = "disagree"
s_d = "strongly disagree"
confidence_num = {"Strongly Agree": 0, "Agree": 0, "Neither Agree Nor Disagree": 0, "Disagree": 0, "Strongly Disagree": 0}
confidence_pct = {"Strongly Agree": 0, "Agree": 0, "Neither Agree Nor Disagree": 0, "Disagree": 0, "Strongly Disagree": 0}

total_so_responses = 0
so_num = {"Heterosexual" : 0, "Homosexual/Gay" : 0, "Bisexual" : 0, "Other": 0, "Prefer Not to Answer" : 0}
so_pct = {"Heterosexual" : 0, "Homosexual/Gay" : 0, "Bisexual" : 0, "Other": 0, "Prefer Not to Answer" : 0}
so_options = {"Heterosexual" , "Homosexual/Gay", "Bisexual", "Prefer Not to Answer"}
other_so_responses = []

total_aa_responses = 0
aa_both = "Both the Perforated Package with Tabs and the Multi-Layer Lubricated Condom"
aa_package_with_tabs = "Perforated Package with Tabs that Allows Condom Application"
aa_condom = "Multi-Layer Lubricated Condom"
aa_num = {aa_both: 0, aa_package_with_tabs: 0, aa_condom : 0, "Neither" : 0}
aa_pct = {aa_both: 0, aa_package_with_tabs: 0, aa_condom : 0, "Neither" : 0}
aa_options = {aa_both, aa_package_with_tabs, aa_condom, "Neither"}

total_sf_responses = 0
stpy = "Several times per year"
stpm = "Several times per month"
stpw = "Several times per week"
sf_options = {stpy, stpm, stpw, "Never"}
sf_num = {stpy : 0, stpm : 0, stpw : 0, "Never": 0}
sf_pct = {stpy : 0, stpm : 0, stpw : 0, "Never": 0}

completeness_num = {"Complete" : 0, "Incomplete" : 0}
completeness_pct = {"Complete" : 0, "Incomplete" : 0}

total_cuf_responses = 0
everyt = "Every time"
halft = "About half of the time"
mhalft = "More than half of the time"
lhalft = "Less than half of the time"
cuf_options = {everyt, halft, mhalft, lhalft, "Never"}
cuf_num = {everyt : 0, mhalft : 0, halft : 0, lhalft : 0, "Never" : 0}
cuf_pct = {everyt : 0, mhalft : 0, halft : 0, lhalft : 0, "Never" : 0}

total_asp_responses = 0
asp_options = {"One", "2 to 5", "5-10", "More than 10"}
asp_num = {"One" : 0, "2 to 5" : 0, "5-10" : 0, "More than 10" : 0}
asp_pct = {"One" : 0, "2 to 5" : 0, "5-10" : 0, "More than 10" : 0}

total_ppc_responses = 0
ppc_num = {"Yes" : 0, "No" : 0}
ppc_pct = {"Yes" : 0, "No" : 0}

total_sti_responses = 0
sti_num = {"Yes" : 0, "No" : 0}
sti_pct = {"Yes" : 0, "No" : 0}

total_lucnp_responses = 0
likely_lucnp = "I would likely use a condom when I have sex with a new partner."
not_sure_lucnp = "I am not sure I would use a condom when I have sex with a new partner."
not_likely_lucnp = "I would likely NOT use a condom when having sex with a new partner."
lucnp_options = {likely_lucnp, not_likely_lucnp, not_sure_lucnp}
lucnp_num = {likely_lucnp : 0, not_likely_lucnp : 0, not_sure_lucnp : 0}
lucnp_pct = {likely_lucnp : 0, not_likely_lucnp : 0, not_sure_lucnp : 0}

total_poce_responses = 0
poce_num = {"Strongly Agree": 0, "Agree": 0, "Neither Agree Nor Disagree": 0, "Disagree": 0, "Strongly Disagree": 0}
poce_pct = {"Strongly Agree": 0, "Agree": 0, "Neither Agree Nor Disagree": 0, "Disagree": 0, "Strongly Disagree": 0}

total_ecn_responses = 0
ecn_num = dict()
ecn_pct = dict()

total_wnc_responses = 0
wnc_num = {"Other" : 0}
wnc_pct = dict()
other_wnc = []

total_ped_reponses = 0
ped_num = dict()
ped_pct = dict()

total_itp1_responses = 0
itp1_num = dict()
itp1_pct = dict()

total_rfi1_responses = 0
rfi1_num = {"Other" : 0}
rfi1_pct = dict()
rfi1_other = []

total_rfn1_responses = 0
rfn1_num = {"Other" : 0}
rfn1_pct = dict()
rfn1_other = []

total_itp2_responses = 0
itp2_num = dict()
itp2_pct = dict()

total_rfi2_responses = 0
rfi2_num = {"Other" : 0}
rfi2_pct = dict()
rfi2_other = []

total_rfn2_responses = 0
rfn2_num = {"Other" : 0}
rfn2_pct = dict()
rfn2_other = []

total_age_responses = 0
age_num = dict()
age_pct = dict()

total_country_responses = 0
country_num = dict()
country_pct = dict()

total_rs_reponses = 0
rs_num = {"Other" : 0}
rs_pct = dict()
rs_other = []

total_gi_responses = 0
gi_num = dict()
gi_pct = dict()

total_trans_responses = 0
trans_num = dict()
trans_pct = dict()

total_race_responses = 0
race_num = {"Other" : 0}
race_pct = dict()
race_other = []

total_econ_responses = 0
econ_num = dict()
econ_pct = dict()

total_god_responses = 0
god_num = dict()
god_pct = dict()

total_dis_responses = 0
dis_yes = 0
dis_dict = {"Other" : 0}
dis_dict_pct = dict()
dis_num = dict()
dis_pct = dict()
dis_other = []

weird = []

master = {"completeness" : dict(), "sf" : dict(), "cuf": dict(), "asp" : dict(), "ppc" : {"Yes" : set(), "No" : set()}, 
            "sti" : {"Yes" : set(), "No" : set()}, "confidence" : dict(), "lucnp" : dict(), "poce" : dict(), 
            "ecn" : dict(), "wnc" : dict(), "ped" : dict(), "itp1" : dict(), "rfi1" : dict(), 
            "rfn1" : dict(), "itp2" : dict(),  "rfi2" : dict(), "rfn2" : dict(), "aa" : dict(), "explainwhy" : dict(), "age" : dict(), 
            "country" : dict(), "rs" : dict(), "gi" : dict(), "trans" : dict(), "so" : dict(), "race" : dict(), 
            "econ" : dict(), "god" : dict(), "dis" : dict(), "disabilities" : dict()}

master_list = []


while data["A" + str(row)].value != None:

    total_responses += 1

    current = Response(row)

    current.setid(data[ "A" + str(row)].value)

    master_list.append(current)



    '''
    BEGIN confidence in using condoms 
    variables:  total_confidence_responses: int, total number of responses with something in condom confidence field
                confidence_num: dict, options (strongly agree, agree, etc.) : number that answered that 
                confidence_pct: dict, options : percentage that answered (out of total_confidence_response)
    '''

    if data[ "I" + str(row)].value != "":
        total_confidence_responses += 1

        if data[ "I" + str(row)].value.lower() == s_a:
            confidence_num["Strongly Agree"] += 1
            current.setconfidence("Strongly Agree")
            if "Strongly Agree" in master["confidence"].keys():
                master["confidence"]["Strongly Agree"].add(current)
            else:
                master["confidence"]["Strongly Agree"] = {current}

        elif data[ "I" + str(row)].value.lower() == _a_:
            confidence_num["Agree"] += 1
            current.setconfidence("Agree")
            if "Agree" in master["confidence"].keys():
                master["confidence"]["Agree"].add(current)
            else:
                master["confidence"]["Agree"] = {current}

        elif data[ "I" + str(row)].value.lower() == n_a_d:
            confidence_num["Neither Agree Nor Disagree"] += 1  
            current.setconfidence("Neither Agree Nor Disagree")
            if "Neither Agree Nor Disagree" in master["confidence"].keys():
                master["confidence"]["Neither Agree Nor Disagree"].add(current)
            else:
                master["confidence"]["Neither Agree Nor Disagree"] = {current}

        elif data[ "I" + str(row)].value.lower() == _d_:
            confidence_num["Disagree"] += 1
            current.setconfidence("Disagree")
            if "Disagree" in master["confidence"].keys():
                master["confidence"]["Disagree"].add(current)
            else:
                master["confidence"]["Disagree"] = {current}


        elif data[ "I" + str(row)].value.lower() == s_d:
            confidence_num["Strongly Disagree"] += 1
            current.setconfidence("Strongly Disagree")
            if "Strongly Disagree" in master["confidence"].keys():
                master["confidence"]["Strongly Disagree"].add(current)
            else:
                master["confidence"]["Strongly Disagree"] = {current}

    '''
    END confidence in using condoms
    '''


    '''
    BEGIN sexual orientation
    variables:  total_so_responses : int, number of responses with something in sexual orientation field
                so_num : dict, orientations [hetero, homo, bi, prefer not to say, other] : number that answered that
                so_pct: dict, orientations : percentage that answered that (out of total_so_responses)
                so_options: set, all the orientation options + other + prefer not to say
                other_so_responses: list of 'other' sexual orientation answers
    '''

    if data[ "AB" + str(row)].value != "":
        total_so_responses += 1

        if data[ "AB" + str(row)].value in so_options:
            so_num[data[ "AB" + str(row)].value] += 1
            current.setso(data[ "AB" + str(row)].value)
            if data[ "AB" + str(row)].value in master["so"].keys():
                master["so"][data[ "AB" + str(row)].value].add(current)
            else:
                master["so"][data[ "AB" + str(row)].value] = {current}

        elif "Straight" in data[ "AB" + str(row)].value or "Normal" in data[ "AB" + str(row)].value:
            so_num["Heterosexual"] += 1
            current.setso("Heterosexual")
            if "Heterosexual" in master["so"].keys():
                master["so"]["Heterosexual"].add(current)
            else:
                master["so"]["Heterosexual"] = {current}

        else:
            so_num["Other"] += 1
            other_so_responses.append(data[ "AB" + str(row)].value)
            current.setso("Other")
            if "Other" in master["so"].keys():
                master["so"]["Other"].add(current)
            else:
                master["so"]["Other"] = {current}



    '''
    END sexual orientation
    '''


    '''
    BEGIN appealing approach
    variables:  total_aa_responses: int, number of responses with something in appealing approach field
                aa_num: dict, options [both, condom, package, neither] : number that answered that
                aa_pct: dict, options : percentage that answered that (out of total_aa_responses)
                aa_options: set of all the appealing approach options 
    '''

    if data[ "U" + str(row)].value != "":
        total_aa_responses += 1

        if data[ "U" + str(row)].value in aa_options:
            aa_num[data[ "U" + str(row)].value] += 1
            current.setaa(data[ "U" + str(row)].value)
            if data[ "U" + str(row)].value in master["aa"].keys():
                master["aa"][data[ "U" + str(row)].value].add(current)
            else:
                master["aa"][data[ "U" + str(row)].value] = {current}
        else:
            weird.append(data[ "U" + str(row)].value)

    '''
    END appealing approach
    '''


    '''
    BEGIN sexual frequency
    variables:  total_sf_responses: int, number of responses with something in sexual frequency field
                sf_num: dict, options [stpy, stpm, stpw, never] : number that answered that
                sf_pct: dict, options : percentage that answered that (out of total_sf_responses)
                sf_options: set, sexual frequency options
    '''

    if data[ "D" + str(row)].value != "":
        total_sf_responses += 1

        if data[ "D" + str(row)].value in sf_options:
            sf_num[data[ "D" + str(row)].value] += 1
            current.setsf(data[ "D" + str(row)].value)
            if data[ "D" + str(row)].value in master["sf"].keys():
                master["sf"][data[ "D" + str(row)].value].add(current)
            else:
                master["sf"][data[ "D" + str(row)].value] = {current}
        else:
            weird.append(data[ "D" + str(row)].value)

    '''
    END sexual frequency
    '''


    '''
    BEGIN completeness
    variables:  completeness_num : dict, complete/incomplete : number of each
                completeness_pct : dict, complete/incomplete : pct of each (out of total_responses)
    '''

    if data[ "B" + str(row)].value == "Complete" or data[ "B" + str(row)].value == "Incomplete":
        completeness_num[data[ "B" + str(row)].value] += 1
        current.setcompleteness(data[ "B" + str(row)].value)
        if data[ "B" + str(row)].value in master["completeness"].keys():
            master["completeness"][data[ "B" + str(row)].value].add(current)
        else:
            master["completeness"][data[ "B" + str(row)].value] = {current}

    '''
    END completeness
    '''


    '''
    BEGIN condom use frequency
    variables:  total_cuf_responses: int, total number of responses with something in condom use frequency field
                cuf_options: set, all cuf options, (always, more than half, half, less than half, never)
                cuf_num: dict, cuf options : number that answered that
                cuf_pct: dict, cuf options : percentage that answered that (out of total_cuf_responses)
    '''

    if data[ "E" + str(row)].value != "":
        total_cuf_responses += 1

        if data[ "E" + str(row)].value in cuf_options:
            cuf_num[data[ "E" + str(row)].value] += 1   
            current.setcuf(data[ "E" + str(row)].value)
            if data[ "E" + str(row)].value in master["cuf"].keys():
                master["cuf"][data[ "E" + str(row)].value].add(current)
            else:
                master["cuf"][data[ "E" + str(row)].value] = {current}

        else:
            weird.append(data[ "E" + str(row)].value)

    '''
    END condom use frequency
    '''


    '''
    BEGIN amount sexual partners
    variables:  total_asp_responses: int, total number of responses with something in amount sexual partners field
                asp_options: set, all asp options (1, 2-5, 5-10, >10)
                asp_num: dict, asp options : number that answered that
                asp_pct: dict, asp options : percentage that answered that (out of total_asp_responses)
    '''

    if data[ "F" + str(row)].value != "":
        total_asp_responses += 1

        if data[ "F" + str(row)].value in asp_options:
            asp_num[data[ "F" + str(row)].value] += 1
            current.setasp(data[ "F" + str(row)].value)
            if data[ "F" + str(row)].value in master["asp"].keys():
                master["asp"][data[ "F" + str(row)].value].add(current)
            else:
                master["asp"][data[ "F" + str(row)].value] = {current}

        else:
            weird.append(data[ "F" + str(row)].value)

    '''
    END amount sexual partners
    '''


    '''
    BEGIN pregnancy prevention concern
    variables:  total_ppc_responses: int, total number of all responses with something in pregnancy prevention concern field
                ppc_num = yes/no : amount that answered that
                ppc_pct = yes/no : percentage that answered that (out of total_ppc_responses)
    '''

    if data[ "G" + str(row)].value != "":
        total_ppc_responses += 1
        ppc_num[data[ "G" + str(row)].value] += 1
        current.setppc(data[ "G" + str(row)].value)
        master["ppc"][data[ "G" + str(row)].value].add(current)

    '''
    END sti prevention concern
    '''

    '''
    BEGIN STI prevention concern
    variables:  sti_
    '''

    if data[ "H" + str(row)].value != "":
        total_sti_responses += 1
        sti_num[data[ "H" + str(row)].value] += 1
        current.setsti(data[ "H" + str(row)].value)
        master["sti"][data[ "H" + str(row)].value].add(current)


    '''
    END STI prevention concern
    '''


    '''
    BEGIN likely use condom new partner
    variables:  lucnp_
    '''

    if data[ "J" + str(row)].value != "":
        total_lucnp_responses += 1
        if data[ "J" + str(row)].value in lucnp_options:
            lucnp_num[data[ "J" + str(row)].value] += 1
            current.setlucnp(data[ "J" + str(row)].value)
            if data[ "J" + str(row)].value in master["lucnp"].keys():
                master["lucnp"][data[ "J" + str(row)].value].add(current)
            else:
                master["lucnp"][data[ "J" + str(row)].value] = {current}    

        else:
            weird.append(data[ "J" + str(row)].value)

    '''
    END likely use condom new partner
    '''


    '''
    BEGIN putting on condom evaluation
    variables: poce_
    '''


    if data[ "K" + str(row)].value != "":
        total_poce_responses += 1

        if data[ "K" + str(row)].value.lower() == s_a:
            poce_num["Strongly Agree"] += 1
            current.setpoce("Strongly Agree")
            if "Strongly Agree" in master["poce"].keys():
                master["poce"]["Strongly Agree"].add(current)
            else:
                master["poce"]["Strongly Agree"] = {current}

        elif data[ "K" + str(row)].value.lower() == _a_:
            poce_num["Agree"] += 1
            current.setpoce("Agree")
            if "Agree" in master["poce"].keys():
                master["poce"]["Agree"].add(current)
            else:
                master["poce"]["Agree"] = {current}

        elif data[ "K" + str(row)].value.lower() == n_a_d:
            poce_num["Neither Agree Nor Disagree"] += 1    
            current.setpoce("Neither Agree Nor Disagree")
            if "Neither Agree Nor Disagree" in master["poce"].keys():
                master["poce"]["Neither Agree Nor Disagree"].add(current)
            else:
                master["poce"]["Neither Agree Nor Disagree"] = {current}

        elif data[ "K" + str(row)].value.lower() == _d_:
            poce_num["Disagree"] += 1
            current.setpoce("Disagree")
            if "Disagree" in master["poce"].keys():
                master["poce"]["Disagree"].add(current)
            else:
                master["poce"]["Disagree"] = {current}
            
        elif data[ "K" + str(row)].value.lower() == s_d:
            poce_num["Strongly Disagree"] += 1 
            current.setpoce("Strongly Disagree")
            if "Strongly Disagree" in master["poce"].keys():
                master["poce"]["Strongly Disagree"].add(current)
            else:
                master["poce"]["Strongly Disagree"] = {current}


    '''
    END putting on condom evaluation
    '''


    '''
    BEGIN easy condom negotiation
    variables:  ecn_
    '''

    if data[ "L" + str(row)].value != "":

        total_ecn_responses += 1
        current.setecn(data[ "L" + str(row)].value)

        if data[ "L" + str(row)].value in ecn_num.keys():
            ecn_num[data[ "L" + str(row)].value] += 1
        else:
            ecn_num[data[ "L" + str(row)].value] = 1
    
        if data[ "L" + str(row)].value in master["ecn"].keys():
            master["ecn"][data[ "L" + str(row)].value].add(current)
        else:
            master["ecn"][data[ "L" + str(row)].value] = {current}

    '''
    END easy condom negotiation
    '''

    '''
    BEGIN why no condom
    variables:  wnc_
                includes other_wnc: list of all the other responses
    '''

    if data[ "M" + str(row)].value != "":

        total_wnc_responses += 1

        responses = data[ "M" + str(row)].value.split("| ")
        master["wnc"]["Other"] = set()

        for response in responses:
            if "Other:" in response:
                master["wnc"]["Other"].add(current)
                other_wnc.append(response)
                wnc_num["Other"] += 1
                current.setwnc("Other")
            elif response not in wnc_num.keys():
                if response in master["wnc"].keys():
                    master["wnc"][response].add(current)
                else:
                    master["wnc"][response] = {current}
                wnc_num[response] = 1
                current.setwnc(response)
            else:
                wnc_num[response] += 1
                if response in master["wnc"].keys():
                    master["wnc"][response].add(current)
                else:
                    master["wnc"][response] = {current}
                current.setwnc(response)
    '''
    END why no condom
    '''

    '''
    BEGIN pleasure-enhancing device evaluation
    variables:  ped_
    '''

    if data[ "N" + str(row)].value != "":
        total_ped_reponses += 1
        current.setped(data[ "N" + str(row)].value)

        if data[ "N" + str(row)].value in ped_num.keys():
            ped_num[data[ "N" + str(row)].value] += 1
            master["ped"][data[ "N" + str(row)].value].add(current)

        else:
            ped_num[data[ "N" + str(row)].value] = 1
            master["ped"][data[ "N" + str(row)].value] = {current}

    '''
    END pleasure-enhancing device evaluation
    '''

    '''
    BEGIN interest in trying product (first time) - corresponds to applicator
    variables:  itp1_
    '''

    if data[ "O" + str(row)].value != "":
        total_itp1_responses += 1
        current.setitp1(data[ "O" + str(row)].value)

        if data[ "O" + str(row)].value in itp1_num.keys():
            itp1_num[data[ "O" + str(row)].value] += 1 
            master["itp1"][data[ "O" + str(row)].value].add(current)
            
        else:
            itp1_num[data[ "O" + str(row)].value] = 1
            master["itp1"][data[ "O" + str(row)].value] = {current}
   
    '''
    END interest in trying product - applicator 
    '''

    '''
    BEGIN reasons for interest (first time) - corresponds to applicator
    variables:  rfi1_
                includes rfi1_other
    '''

    if data[ "P" + str(row)].value != "":
        total_rfi1_responses += 1
        master["rfi1"]["Other"] = set()

        responses = data[ "P" + str(row)].value.split("| ")
        for response in responses:
            if "Other:" in response:
                master["rfi1"]["Other"].add(current)
                rfi1_other.append(response)
                rfi1_num["Other"] += 1
                current.setrfi1("Other")
            elif response in rfi1_num.keys():
                rfi1_num[response] += 1
                master["rfi1"][response].add(current)
                current.setrfi1(response)
            else:
                master["rfi1"][response] = {current}
                rfi1_num[response] = 1
                current.setrfi1(response)

    '''
    END reasons for interest - applicator
    '''

    '''
    BEGIN reasons for no interest (first time) - corresponds to applicator 
    variables:  rfn1_
                includes rf1n_other
    '''

    if data[ "Q" + str(row)].value != "":
        total_rfn1_responses += 1
        master["rfn1"]["Other"] = set()

        responses = data[ "Q" + str(row)].value.split("| ")
        for response in responses:
            if "Other:" in response:
                master["rfn1"]["Other"].add(current)
                rfn1_other.append(response)
                rfn1_num["Other"] += 1
                current.setfrn1("Other")
            elif response in rfn1_num.keys():
                rfn1_num[response] += 1
                master["rfn1"][response].add(current)
                current.setfrn1(response)
            else:
                rfn1_num[response] = 1
                master["rfn1"][response] = {current}
                current.setfrn1(response)

    '''
    END reasons for no interest - applicator
    '''

    '''
    BEGIN interest in trying product (second time) - corresponds to multi-layer condom
    variables:  itp2_
    '''

    if data[ "R" + str(row)].value != "":
        total_itp2_responses += 1
        current.setitp2(data[ "R" + str(row)].value)
        

        if data[ "R" + str(row)].value in itp2_num.keys():
            itp2_num[data[ "R" + str(row)].value] += 1 
            master["itp2"][data[ "R" + str(row)].value].add(current)
        else:
            itp2_num[data[ "R" + str(row)].value] = 1
            master["itp2"][data[ "R" + str(row)].value] = {current}


    '''
    END interest in trying product - mlc
    '''
    '''
    BEGIN reasons for interest (second time) - corresponds to multi-layer condom
    variables:  rfi2_
                includes rfi2_other
    '''

    if data[ "S" + str(row)].value != "":
        total_rfi2_responses += 1
        master["rfi2"]["Other"] = set()

        responses = data[ "S" + str(row)].value.split("| ")
        for response in responses:
            current.setrfi2(response)
            if "Other:" in response:
                master["rfi2"]["Other"].add(current)
                rfi2_other.append(response)
                rfi2_num["Other"] += 1
            elif response in rfi2_num.keys():
                rfi2_num[response] += 1
                master["rfi2"][response].add(current)
            else:
                rfi2_num[response] = 1
                master["rfi2"][response] = {current}
                
    '''
    END reasons for interest - mlc
    '''

    '''
    BEGIN reasons for no interest (second time) - corresponds to multi-layer condom 
    variables:  rfn2_
                includes rfn2_other
    '''

    if data[ "T" + str(row)].value != "":
        total_rfn2_responses += 1
        master["rfn2"]["Other"] = set()
        
        responses = data[ "T" + str(row)].value.split("| ")
        for response in responses:
            current.setrfn2(response)
            if "Other:" in response:
                rfn2_other.append(response)
                rfn2_num["Other"] += 1
                master["rfn2"]["Other"].add(current)
            elif response in rfn2_num.keys():
                rfn2_num[response] += 1
                master["rfn2"][response].add(current)
            else:
                rfn2_num[response] = 1
                master["rfn2"][response] = {current}


    '''
    END reasons for no interest - mlc
    '''

    '''
    BEGIN age
    variables:  age_
    '''

    if data[ "W" + str(row)].value != "":
        total_age_responses += 1
        current.setage(data[ "W" + str(row)].value)

        if data[ "W" + str(row)].value in age_num.keys():
            age_num[data[ "W" + str(row)].value] += 1
            master["age"][data[ "W" + str(row)].value].add(current)
        else:
            age_num[data[ "W" + str(row)].value] = 1
            master["age"][data[ "W" + str(row)].value] = {current}
    

    '''
    END age
    '''

    '''
    BEGIN country
    variables:  country_
    '''

    if data[ "X" + str(row)].value != "":
        total_country_responses += 1
        current.setcountry(data[ "X" + str(row)].value)

        if data[ "X" + str(row)].value in country_num.keys():
            country_num[data[ "X" + str(row)].value] += 1
            master["country"][data[ "X" + str(row)].value].add(current)
        else:
            country_num[data[ "X" + str(row)].value] = 1
            master["country"][data[ "X" + str(row)].value] = {current}

    '''
    END country
    '''

    '''
    BEGIN relationship status
    variables:  rs_
                includes rs_other
    '''

    if data[ "Y" + str(row)].value != "":
        total_rs_reponses += 1
        master["rs"]["Other"] = set()

        if "Other:" in data[ "Y" + str(row)].value:
            rs_num["Other"] += 1
            rs_other.append(data[ "Y" + str(row)].value)
            master["rs"]["Other"].add(current)
            current.setrs("Other")
        elif data[ "Y" + str(row)].value in rs_num.keys():
            rs_num[data[ "Y" + str(row)].value] += 1
            master["rs"][data[ "Y" + str(row)].value].add(current)
            current.setrs(data[ "Y" + str(row)].value)
        else:
            rs_num[data[ "Y" + str(row)].value] = 1
            master["rs"][data[ "Y" + str(row)].value] = {current}
            current.setrs(data[ "Y" + str(row)].value)
    
    '''
    END relationship status
    '''

    '''
    BEGIN gender identity
    variables:  gi_
    '''

    if data[ "Z" + str(row)].value != "":
        total_gi_responses += 1
        current.setgi(data[ "Z" + str(row)].value)

        if data[ "Z" + str(row)].value in gi_num.keys():
            gi_num[data[ "Z" + str(row)].value] += 1
            master["gi"][data[ "Z" + str(row)].value].add(current)
        else:
            gi_num[data[ "Z" + str(row)].value] = 1
            master["gi"][data[ "Z" + str(row)].value] = {current}

    '''
    END gender identity
    '''

    '''
    BEGIN transgender
    variables:  trans_
    '''


    if data[ "AA" + str(row)].value != "":
        total_trans_responses += 1
        current.settrans(data[ "AA" + str(row)].value)

        if data[ "AA" + str(row)].value in trans_num.keys():
            trans_num[data[ "AA" + str(row)].value] += 1
            master["trans"][data[ "AA" + str(row)].value].add(current)
        else:
            trans_num[data[ "AA" + str(row)].value] = 1
            master["trans"][data[ "AA" + str(row)].value] = {current}
        
    '''
    END transgender
    '''

    '''
    BEGIN racial/ethnic identity
    variables:  race_
                includes race_other
    '''

    if data[ "AC" + str(row)].value != "":
        total_race_responses += 1
        master["race"]["Other"] = set()
        responses = data[ "AC" + str(row)].value.split("| ")

        for response in responses:
            if "Other:" in response:
                race_num["Other"] += 1
                race_other.append(response)
                master["race"]["Other"].add(current)
                current.setrace("Other")
            elif response in race_num.keys():
                race_num[response] += 1
                master["race"][response].add(current)
                current.setrace(response)
            else:
                race_num[response] = 1
                master["race"][response] = {current}
                current.setrace(response)

    '''
    END racial/ethnic identity
    '''


    '''
    BEGIN economic situation
    variables:  econ_
    '''

    if data[ "AD" + str(row)].value != "":
        total_econ_responses += 1
        current.setecon(data[ "AD" + str(row)].value)

        if data[ "AD" + str(row)].value in econ_num.keys():
            econ_num[data[ "AD" + str(row)].value] += 1
            master["econ"][data[ "AD" + str(row)].value].add(current)

        else:
            econ_num[data[ "AD" + str(row)].value] = 1
            master["econ"][data[ "AD" + str(row)].value] = {current}

    '''
    END economic situation
    '''


    '''
    BEGIN religious beliefs
    variables:  god_
    '''

    if data[ "AE" + str(row)].value != "":
        total_god_responses += 1
        current.setgod(data[ "AE" + str(row)].value)

        if data[ "AE" + str(row)].value in god_num.keys():
            god_num[data[ "AE" + str(row)].value] += 1
            master["god"][data[ "AE" + str(row)].value].add(current)
        else:
            god_num[data[ "AE" + str(row)].value] = 1
            master["god"][data[ "AE" + str(row)].value] = {current}
    
    '''
    END religious beliefs
    '''
    
    '''
    BEGIN disabilities
    variables:  dis_ 
                includes    dis_dict, a dict of the frequency of each disability
                            dis_dict_pct, a pct dictionary for each disability, out of dis_yes
                            dis_yes: int, number of people who responded yes
                            dis_other: list of 'other' disabilities mentioned
    '''

    if data[ "AF" + str(row)].value != "":
        total_dis_responses += 1
        current.setdis(data[ "AF" + str(row)].value)

        if data[ "AF" + str(row)].value in dis_num.keys():
            dis_num[data[ "AF" + str(row)].value] += 1
            master["dis"][data[ "AF" + str(row)].value].add(current)
        else:
            dis_num[data[ "AF" + str(row)].value] = 1
            master["dis"][data[ "AF" + str(row)].value] = {current}
        
        master["disabilities"]["Other"] = set()
        if data[ "AF" + str(row)].value == "Yes":
            dis_yes += 1
            list_dis = data[ "AG" + str(row)].value.split("| ")
            for dis in list_dis:
                if "Other:" in dis:
                    dis_other.append(dis[7:])
                    dis_dict["Other"] += 1
                    master["disabilities"]["Other"].add(current)
                    current.setdisabilites("Other")
                elif dis in dis_dict.keys():
                    master["disabilities"][dis].add(current)
                    dis_dict[dis] += 1
                    current.setdisabilites(dis)
                else:
                    dis_dict[dis] = 1
                    master["disabilities"][dis] = {current}
                    current.setdisabilites(dis)




    row += 1


    
confidence_pct["Strongly Agree"] = confidence_num["Strongly Agree"] / total_confidence_responses
confidence_pct["Agree"] = confidence_num["Agree"] / total_confidence_responses
confidence_pct["Neither Agree Nor Disagree"] = confidence_num["Neither Agree Nor Disagree"] / total_confidence_responses
confidence_pct["Disagree"] = confidence_num["Disagree"] / total_confidence_responses
confidence_pct["Strongly Disagree"] = confidence_num["Strongly Disagree"] / total_confidence_responses

so_pct["Heterosexual"] = so_num["Heterosexual"] / total_so_responses
so_pct["Homosexual/Gay"] = so_num["Homosexual/Gay"] / total_so_responses
so_pct["Bisexual"] = so_num["Bisexual"] / total_so_responses
so_pct["Other"] = so_num["Other"] / total_so_responses
so_pct["Prefer Not to Answer"] = so_num["Prefer Not to Answer"] / total_so_responses

aa_pct[aa_both] = aa_num[aa_both] / total_aa_responses
aa_pct[aa_package_with_tabs] = aa_num[aa_package_with_tabs] / total_aa_responses
aa_pct[aa_condom] = aa_num[aa_condom] / total_aa_responses
aa_pct["Neither"] = aa_num["Neither"] / total_aa_responses

sf_pct[stpy] = sf_num[stpy] / total_sf_responses
sf_pct[stpm] = sf_num[stpm] / total_sf_responses
sf_pct[stpw] = sf_num[stpw] / total_sf_responses
sf_pct["Never"] = sf_num["Never"] / total_sf_responses

completeness_pct["Incomplete"] = completeness_num["Incomplete"] / total_responses
completeness_pct["Complete"] = completeness_num["Complete"] / total_responses

for option in cuf_options:
    cuf_pct[option] = cuf_num[option] / total_cuf_responses

for option in asp_options:
    asp_pct[option] = asp_num[option] / total_asp_responses

ppc_pct["Yes"] = ppc_num["Yes"] / total_ppc_responses
ppc_pct["No"] = ppc_num["No"] / total_ppc_responses

sti_pct["Yes"] = sti_num["Yes"] / total_sti_responses
sti_pct["No"] = sti_num["No"] / total_sti_responses

for option in lucnp_options:
    lucnp_pct[option] = lucnp_num[option] / total_lucnp_responses

for option in poce_num:
    poce_pct[option] = poce_num[option] / total_poce_responses

for option in ecn_num:
    ecn_pct[option] = ecn_num[option] / total_ecn_responses

for option in wnc_num:
    wnc_pct[option] = wnc_num[option] / total_wnc_responses

for option in ped_num:
    ped_pct[option] = ped_num[option] / total_ped_reponses

for option in itp1_num:
    itp1_pct[option] = itp1_num[option] / total_itp1_responses

for option in rfi1_num:
    rfi1_pct[option] = rfi1_num[option] / total_rfi1_responses

for option in rfn1_num:
    rfn1_pct[option] = rfn1_num[option] / total_rfn1_responses

for option in itp2_num:
    itp2_pct[option] = itp2_num[option] / total_itp2_responses

for option in rfi2_num:
    rfi2_pct[option] = rfi2_num[option] / total_rfi2_responses

for option in rfn2_num:
    rfn2_pct[option] = rfn2_num[option] / total_rfn2_responses

for option in age_num:
    age_pct[option] = age_num[option] / total_age_responses

for option in country_num:
    country_pct[option] = country_num[option] / total_country_responses

for option in rs_num:
    rs_pct[option] = rs_num[option] / total_rs_reponses

for option in gi_num:
    gi_pct[option] = gi_num[option] / total_gi_responses

for option in trans_num:
    trans_pct[option] = trans_num[option] / total_trans_responses

for option in race_num:
    race_pct[option] = race_num[option] / total_race_responses

for option in econ_num:
    econ_pct[option] = econ_num[option] / total_race_responses

for option in god_num:
    god_pct[option] = god_num[option] / total_god_responses

for option in dis_num:
    dis_pct[option] = dis_num[option] / total_dis_responses

for dis in dis_dict:
    dis_dict_pct[dis] = dis_dict[dis] / dis_yes


if __name__ == "__main__":

    # print (total_responses)

    # print (total_confidence_responses)
    # print (confidence_num)
    # print (confidence_pct)

    # print (total_so_responses)
    # print (so_num)
    # print (so_pct)
    # print (other_so_responses)

    # print (total_aa_responses)
    # print (aa_num)
    # print (aa_pct)

    # print (total_sf_responses)
    # print (sf_num)
    # print (sf_pct)

    # print (total_responses)
    # print (completeness_num)
    # print (completeness_pct)

    # print (total_cuf_responses)
    # print (cuf_num)
    # print (cuf_pct)

    # print (total_asp_responses)
    # print (asp_num)
    # print (asp_pct)

    # print (total_ppc_responses)
    # print (ppc_num)
    # print (ppc_pct)

    # print (total_sti_responses)
    # print (sti_num)
    # print (sti_pct)

    # print (total_lucnp_responses)
    # print (lucnp_num)
    # print (lucnp_pct)

    # print (total_poce_responses)
    # print (poce_num)
    # print (poce_pct)

    # print (lucnp_num)
    # print (lucnp_pct)

    # print(total_ecn_responses)
    # print (ecn_num)
    # print (ecn_pct)

    # print (total_wnc_responses)
    # print (wnc_num)
    # print (wnc_pct)
    # print (len(other_wnc))

    # print (total_ped_reponses)
    # print (ped_num)
    # print (ped_pct)

    # print (total_itp1_responses)
    # print (itp1_num)
    # print (itp1_pct)

    # print (total_itp2_responses)
    # print (itp2_num)
    # print (itp2_pct)

    # print (total_rfi1_responses)
    # print (rfi1_num)
    # print (rfi1_pct)
    
    # print (total_rfn1_responses)
    # print (rfn1_num)
    # print (rfn1_pct)

    # print (total_rfi2_responses)

    # print (total_rfn2_responses)
    # print (rfn2_num)
    # print (rfn2_pct)

    # print (total_age_responses)
    # print (age_num)
    # print (age_pct)

    # print (total_country_responses)
    # print (country_num)
    # print (country_pct)
    # print (len(country_num.keys()))

    # print (total_rs_reponses)
    # print (rs_num)
    # print (rs_pct)
    # print (rs_other)

    # print (total_gi_responses)
    # print (gi_num)
    # print (gi_pct)

    # print (total_trans_responses)
    # print (trans_num)
    # print (trans_pct)

    # print (total_race_responses)
    # print (race_num)
    # print (race_pct)

    # print (econ_num)

    # print (god_num)

    # print (dis_num)
    # print (dis_pct)
    # print (dis_dict)
    # print (dis_dict_pct)
    # print (dis_other)


    #print (master_list[0].confidence)


    # print ("weird:", weird)
    pass

