from typing import Pattern
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles.colors import Color
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

import os



datawb = load_workbook("CondomSurvey.xlsx")
data = datawb.active

class Response:

    def __init__(self, row):

        self.row = row
        self.id = None
        self.completeness = None
        self.sf = None
        self.cuf = None
        self.asp = None
        self.ppc = None
        self.sti = None
        self.confidence = None
        self.lucnp = None
        self.poce = None
        self.ecn = None
        self.wnc = []
        self.ped = None
        self.itp1 = None
        self.rfi1 = []
        self.rfn1 = []
        self.itp2 = None
        self.rfi2 = []
        self.rfn2 = []
        self.aa = None
        self.explainwhy = None
        self.age = None
        self.country = None
        self.rs = None
        self.gi = None
        self.trans = None
        self.so = None
        self.race = []
        self.econ = None
        self.god = None
        self.dis = None
        self.disabilities = []

    def setid(self, id1):
        self.id = id1
    
    def setcompleteness(self, completeness):
        self.completeness = completeness

    def setsf(self, sf):
        self.sf = sf

    def setcuf(self, cuf):
        self.cuf = cuf

    def setasp(self, asp):
        self.asp = asp

    def setppc(self, ppc):
        self.ppc = ppc
    
    def setsti(self, sti):
        self.sti = sti
    
    def setconfidence(self, confidence):
        self.confidence = confidence
    
    def setlucnp(self, lucnp):
        self.lucnp = lucnp
    
    def setpoce(self, poce):
        self.poce = poce
    
    def setecn(self, ecn):
        self.ecn = ecn
    
    def setwnc(self, wnc):
        self.wnc.append(wnc)
    
    def setped(self, ped):
        self.ped = ped
    
    def setitp1(self, applicatorinterest):
        self.itp1 = applicatorinterest

    def setrfi1(self, applicatorreason):
        self.rfi1.append(applicatorreason)
    
    def setfrn1(self, noapplicatorreason):
        self.rfn1.append(noapplicatorreason)

    def setitp2(self, mlcinterest):
        self.itp2 = mlcinterest

    def setrfi2(self, mlcreason):
        self.rfi2.append(mlcreason)

    def setrfn2(self, nomlcreason):
        self.rfn2.append(nomlcreason)

    def setaa(self, aa):
        self.aa = aa
    
    def setexplainwhy(self, explainwhy):
        self.explainwhy = explainwhy
    
    def setage(self, age):
        self.age = age
    
    def setcountry(self, country):
        self.country = country

    def setrs(self, rs):
        self.rs = rs
    
    def setgi(self,gender):
        self.gi = gender
    
    def settrans(self, transgender):
        self.trans = transgender
    
    def setso(self, so):
        self.so = so
    
    def setrace(self, race):
        self.race.append(race)
    
    def setecon(self, economic):
        self.econ = economic
    
    def setgod(self, god):
        self.god = god
    
    def setdis(self, disabled):
        self.dis = disabled
    
    def setdisabilites(self, disability):
        self.disabilities.append(disability)
















        pass